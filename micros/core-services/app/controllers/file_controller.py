from ..mysql import Database
from ..models.file_model import FileModel
from pathlib import Path
from ..config import get_env
from uuid import uuid4 as uuid
from fastapi import UploadFile
import openpyxl
import json

class FileController:
    
  async def upload_file(self, file: UploadFile, id_inventario: int):
    MAX_FILE_SIZE = get_env().MAX_FILE_SIZE
    ALLOWED_EXTENSIONS = get_env().ALLOWED_EXTENSIONS.split(",")
    id_usuario = 5  # Valor estático
    id_perfil = 2   # Valor estático

    if not id_inventario:
        print(" El id_inventario no es válido.")
        return "ID Inventario no proporcionado o inválido"

    with Database() as db:
        try:
            print("Iniciando carga de archivo...")
            file_content = await file.read()
            file_size = len(file_content)
            print(f"El tamaño del archivo es: {file_size} bytes")

            if file_size > MAX_FILE_SIZE:
                print("El archivo es demasiado grande")
                return "File too large"
            
            file_extension = Path(file.filename).suffix
            print(f"Extensión del archivo: {file_extension}")
            
            if file_extension not in ALLOWED_EXTENSIONS:
                print("El tipo de archivo no es permitido")
                return "File type not allowed"

            # folder_path = "C:\\Users\\pope0\\OneDrive\\Escritorio\\arcivhos_conteo"
            folder_path = "D:\\CRISTIAN\\Escritorio\\templateProyect"
            Path(folder_path).mkdir(parents=True, exist_ok=True)
            file_path = Path(folder_path) / f"{str(uuid())}.{file_extension}"
            with open(file_path, "wb") as buffer:
                buffer.write(file_content)
            print(f"Archivo guardado en: {file_path}")
            
            wb_obj = openpyxl.load_workbook(file_path)
            sheet_obj = wb_obj.active
            wb_obj.close()
            
            headers = [sheet_obj.cell(row=1, column=col).value.strip().upper() for col in range(1, sheet_obj.max_column + 1)]  # Normalizar a mayúsculas
            print(f"Encabezados del archivo: {headers}")

            check_query = "SELECT id_inventario FROM inventarios WHERE id_inventario = %s"
            db.execute(check_query, (id_inventario,))
            if db.fetchone() is None:
                return f"El id inventario {id_inventario} no existe"

            # Verificar y agregar las columnas dinámicas a la tabla 'productos'
            for header in headers:
                if header not in ["CODIGO DE BARRAS", "STOCK"]:  # Estas columnas ya existen
                    try:
                        alter_query = f"ALTER TABLE productos ADD COLUMN IF NOT EXISTS `{header}` VARCHAR(255)"
                        db.execute(alter_query)
                    except Exception as e:
                        print(f"Error al agregar columna {header}: {str(e)}")

            for row in range(2, sheet_obj.max_row + 1):
                print(f"Procesando fila {row}")
                
                data = {}
                row_data = {}
                
                for col_index, header in enumerate(headers, start=1):
                    row_data[header] = sheet_obj.cell(row=row, column=col_index).value
                
                codigo_barras = row_data.get("CODIGO DE BARRAS")
                stock = row_data.get("STOCK")
                
                if codigo_barras is None or stock is None:
                    print(f" Faltan datos en la fila {row}")
                    continue 
                
                for key, value in row_data.items():
                    if key != "STOCK" and key != "CODIGO DE BARRAS":
                        data[key] = value  # Guardar los datos en 'data'

                print("Datos de la fila:", row_data)
                print("Ejecutando INSERT con:", row_data["CODIGO DE BARRAS"], row_data["STOCK"], json.dumps(data))

                query = """
                    INSERT INTO productos (codigo_barras, stock, data, id_inventario, id_usuario, id_perfil)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """
                
                try:
                    db.execute(
                        query,
                        (row_data["CODIGO DE BARRAS"], row_data["STOCK"], json.dumps(data), id_inventario, id_usuario, id_perfil )
                    )
                    db.commit()
                    print(f" Insert realizado con éxito en la fila {row}")
                except Exception as e:
                    db.rollback()
                    print(f" Error al hacer INSERT en la fila {row}: {str(e)}")

            return True
        
        except Exception as e:
            db.rollback()
            print(f" Error en la carga del archivo: {str(e)}")
            return {"success": False, "error": str(e)}


   
    
    
    # async def upload_file(self, file: UploadFile, id_inventario: int):
        
    #     MAX_FILE_SIZE = get_env().MAX_FILE_SIZE
    #     ALLOWED_EXTENSIONS = get_env().ALLOWED_EXTENSIONS.split(",")
    #     id_usuario = 5
    #     id_perfil= 2

    #     if not id_inventario:
    #         print(" El id_inventario no es válido.")
    #         return "ID Inventario no proporcionado o inválido"

    #     with Database() as db:
    #         try:
    #             print("Iniciando carga de archivo...")
    #             file_content = await file.read()
    #             file_size = len(file_content)
    #             print(f"El tamaño del archivo es: {file_size} bytes")

    #             if file_size > MAX_FILE_SIZE:
    #                 print("El archivo es demasiado grande")
    #                 return "File too large"
                
    #             file_extension = Path(file.filename).suffix
    #             print(f"Extensión del archivo: {file_extension}")
                
    #             if file_extension not in ALLOWED_EXTENSIONS:
    #                 print("El tipo de archivo no es permitido")
    #                 return "File type not allowed"

    #             # folder_path = "C:\\Users\\pope0\\OneDrive\\Escritorio\\arcivhos_conteo"
    #             folder_path = "D:\\CRISTIAN\\Escritorio\\templateProyect"
    #             Path(folder_path).mkdir(parents=True, exist_ok=True)
    #             file_path = Path(folder_path) / f"{str(uuid())}.{file_extension}"
    #             with open(file_path, "wb") as buffer:
    #                 buffer.write(file_content)
    #             print(f"Archivo guardado en: {file_path}")
                
    #             wb_obj = openpyxl.load_workbook(file_path)
    #             sheet_obj = wb_obj.active
    #             wb_obj.close()
                
    #             headers = [sheet_obj.cell(row=1, column=col).value.strip().upper() for col in range(1, sheet_obj.max_column + 1)]  # Normalizar a mayúsculas
    #             print(f"Encabezados del archivo: {headers}")

    #             check_query = "SELECT id_inventario FROM inventarios WHERE id_inventario = %s"
    #             db.execute(check_query, (id_inventario,))
    #             if db.fetchone() is None:
    #                 return f"El id inventario {id_inventario} no existe"
                
                

    #             for row in range(2, sheet_obj.max_row + 1):
    #                 print(f"Procesando fila {row}")
                    
    #                 data = {}
    #                 row_data = {}
                    
    #                 for col_index, header in enumerate(headers, start=1):
    #                     row_data[header] = sheet_obj.cell(row=row, column=col_index).value
                    
    #                 codigo_barras = row_data.get("CODIGO DE BARRAS")
    #                 stock = row_data.get("STOCK")
                    
    #                 if codigo_barras is None or stock is None:
    #                     print(f" Faltan datos en la fila {row}")
    #                     continue 
                    
    #                 for key, value in row_data.items():
    #                     if key != "STOCK" and key != "CODIGO DE BARRAS":
    #                         data[key] = value  

    #                 print("Datos de la fila:", row_data)
    #                 print("Ejecutando INSERT con:", row_data["CODIGO DE BARRAS"], row_data["STOCK"], json.dumps(data))


    #                 query = """
    #                     INSERT INTO productos (codigo_barras, stock, data, id_inventario, id_usuario, id_perfil)
    #                     VALUES (%s, %s, %s, %s, %s, %s)
    #                 """
                    
    #                 try:
    #                     db.execute(
    #                         query,
    #                         (row_data["CODIGO DE BARRAS"], row_data["STOCK"], json.dumps(data), id_inventario, id_usuario, id_perfil )
    #                     )
    #                     db.commit()
    #                     print(f" Insert realizado con éxito en la fila {row}")
    #                 except Exception as e:
    #                     db.rollback()
    #                     print(f" Error al hacer INSERT en la fila {row}: {str(e)}")

    #             return True
            
    #         except Exception as e:
    #             db.rollback()
    #             print(f" Error en la carga del archivo: {str(e)}")
    #             return {"success": False, "error": str(e)}






